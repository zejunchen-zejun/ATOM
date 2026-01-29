import os
import sys
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_log_file(log_file_path):
    """Parse log file and extract key metrics"""
    results = []

    with open(log_file_path, "r", encoding="utf-8") as f:
        content = f.read()

    test_blocks = re.findall(
        r"Start Test ISL=(\d+), OSL=(\d+), CONC=(\d+)(.*?)==================================================",
        content,
        re.DOTALL,
    )

    for isl, osl, conc, block in test_blocks:
        mean_ttft = re.search(r"Mean TTFT \(ms\):\s+([\d.]+)", block)
        mean_tpot = re.search(r"Mean TPOT \(ms\):\s+([\d.]+)", block)
        total_throughput = re.search(
            r"Total Token throughput \(tok/s\):\s+([\d.]+)", block
        )

        if mean_ttft and mean_tpot and total_throughput:
            results.append(
                {
                    "ISL": int(isl),
                    "OSL": int(osl),
                    "CONC": int(conc),
                    "Mean TTFT (ms)": float(mean_ttft.group(1)),
                    "Mean TPOT (ms)": float(mean_tpot.group(1)),
                    "Total Throughput (tok/s)": float(total_throughput.group(1)),
                }
            )

    return results


def create_excel_repeating_format(results, output_file):
    """Create Excel file with repeating format: each ISL/OSL combination repeats display, ensure ISL row has complete border"""

    # Sort by ISL and CONC
    results.sort(key=lambda x: (x["ISL"], -x["CONC"]))

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Define border style - ensure all borders are set
    full_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Define fill styles
    header_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    isl_header_fills = [
        PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
    ]

    # Define font styles
    header_font = Font(bold=True, size=11, color="000000")
    isl_title_font = Font(bold=True, size=12, color="000000")
    data_font = Font(size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")

    current_row = 1
    grouped_results = {}
    for result in results:
        key = (result["ISL"], result["OSL"])
        if key not in grouped_results:
            grouped_results[key] = []
        grouped_results[key].append(result)

    # Sort by ISL
    sorted_keys = sorted(grouped_results.keys())

    # Create independent 4-row area for each ISL/OSL combination
    for idx, (isl, osl) in enumerate(sorted_keys):
        data_list = grouped_results[(isl, osl)]
        data_list.sort(key=lambda x: -x["CONC"])  # Sort by CONC descending

        # Row 1: ISL/OSL title row (merge 4 cells) - ensure complete border
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )

        cell = ws.cell(row=current_row, column=1, value=f"ISL={isl}, OSL={osl}")
        cell.fill = isl_header_fills[idx % len(isl_header_fills)]
        cell.font = isl_title_font
        cell.alignment = center_alignment
        cell.border = full_border  # Ensure complete border

        # Set border for each cell of merged cells individually (ensure display)
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = full_border

        current_row += 1

        # Row 2: Header row
        for col, header in enumerate(
            [
                "Concurrency",
                "Mean TTFT (ms)",
                "Mean TPOT (ms)",
                "Total Throughput (tok/s)",
            ],
            1,
        ):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = full_border

        current_row += 1

        # Row 3-4: Data rows (maximum 2 concurrency values)
        for i, data in enumerate(data_list):  # Take first 2 concurrency values
            # Concurrency
            cell = ws.cell(row=current_row, column=1, value=data["CONC"])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border

            # Mean TTFT
            cell = ws.cell(row=current_row, column=2, value=data["Mean TTFT (ms)"])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border

            # Mean TPOT
            cell = ws.cell(row=current_row, column=3, value=data["Mean TPOT (ms)"])
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border

            # Total Throughput
            cell = ws.cell(
                row=current_row, column=4, value=data["Total Throughput (tok/s)"]
            )
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = full_border

            current_row += 1

        # Add empty row to separate different ISL/OSL combinations (no border)
        current_row += 1

    # Remove last empty row
    if current_row > 1:
        current_row -= 1

    # Auto adjust column width
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)

        if col_idx == 1:  # Concurrency column
            ws.column_dimensions[column_letter].width = 12
        elif col_idx == 2 or col_idx == 3:  # Mean TTFT/TPOT columns
            ws.column_dimensions[column_letter].width = 15
        else:  # Total Throughput column
            ws.column_dimensions[column_letter].width = 22

    # Save file
    wb.save(output_file)
    return output_file


def print_csv_repeating_format(results):
    """Print data in CSV table format (repeating format)"""

    # Group by ISL
    grouped_results = {}
    for result in results:
        key = (result["ISL"], result["OSL"])
        if key not in grouped_results:
            grouped_results[key] = []
        grouped_results[key].append(result)

    # Sort by ISL
    sorted_keys = sorted(grouped_results.keys())

    print("\n" + "=" * 60)
    print("DEEPSEEK-V3.2 BENCHMARK RESULTS")
    print("=" * 60)

    # Print independent table for each ISL/OSL combination
    for idx, (isl, osl) in enumerate(sorted_keys):
        data_list = grouped_results[(isl, osl)]
        data_list.sort(key=lambda x: -x["CONC"])  # Sort by CONC descending

        print(f"\n┌{'ISL=' + str(isl) + ', OSL=' + str(osl):^52}┐")
        print("├────────────────────────────────────────────────────┤")
        print("│ Concurrency │ Mean TTFT │ Mean TPOT │  Throughput  │")
        print("├─────────────┼───────────┼───────────┼──────────────┤")

        # Print maximum 2 rows of data
        for i, data in enumerate(data_list):
            print(
                f"│ {data['CONC']:^11} │ {data['Mean TTFT (ms)']:>9.2f} │ {data['Mean TPOT (ms)']:>9.2f} │ {data['Total Throughput (tok/s)']:>12.2f} │"
            )

        # If data has less than 2 rows, fill with empty rows
        while len(data_list) < 2 and i < 1:
            print("│             │           │           │              │")
            i += 1

        print("└─────────────┴───────────┴───────────┴──────────────┘")

    print("=" * 60)


def main():
    # Input file path
    parser = argparse.ArgumentParser(
        description="benchmark infomation",
    )

    # Add parameter
    parser.add_argument(
        "input_file",
        nargs="?",  # Optional parameter
        default="result.txt",
        help="the path of input file, default:result.txt.",
    )

    args = parser.parse_args()
    log_file_path = args.input_file

    if not os.path.isfile(log_file_path):
        print(f"file not exist - {log_file_path}")
        sys.exit(1)

    output_file = "benchmark.xlsx"

    try:
        # Parse log file
        results = parse_log_file(log_file_path)

        if results:
            # Print data in CSV table format
            print_csv_repeating_format(results)

            # Create Excel file
            create_excel_repeating_format(results, output_file)

        else:
            print("No valid benchmark results found")

    except FileNotFoundError:
        print(f"Error: cannot find log file {log_file_path}")
    except Exception as e:
        print(f"Error during processing: {e}")


if __name__ == "__main__":
    main()
